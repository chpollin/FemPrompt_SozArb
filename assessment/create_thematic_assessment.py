#!/usr/bin/env python3
"""
FemPrompt Thematic Assessment Excel Generator

Creates formatted Excel file with 14 thematic columns (Ja/Nein) for paper assessment.
Fetches data from Zotero API and adds:
- Technik-Dimensionen: AI_Literacies, Generative_KI, Prompting, KI_Sonstige
- Sozial-Dimensionen: Soziale_Arbeit, Bias_Ungleichheit, Gender, Diversitaet, Feministisch, Fairness
- Meta: Studientyp, Decision, Exclusion_Reason, Notes

Features:
- Color-coded headers (Technik: blue, Sozial: green, Meta: orange)
- Dropdown validation for Ja/Nein, Decision, Studientyp
- Frozen header row
- Alphabetical sorting by Author_Year
- Abstract whitespace normalization
- Google Sheets compatible

Usage:
    python assessment/create_thematic_assessment.py
    python assessment/create_thematic_assessment.py --output assessment/assessment_new.xlsx
    python assessment/create_thematic_assessment.py --library-id 6080294
"""

import os
import sys
import re
import argparse
import logging
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    from pyzotero import zotero
except ImportError:
    print("Error: pyzotero not installed. Run: pip install pyzotero")
    sys.exit(1)

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ThematicAssessmentGenerator:
    """Generate FemPrompt thematic assessment Excel from Zotero"""

    # Color definitions
    COLORS = {
        'header_meta': 'B4C6E7',      # Light blue - metadata columns
        'header_technik': '4472C4',    # Blue - Technik dimensions
        'header_sozial': '70AD47',     # Green - Sozial dimensions
        'header_assessment': 'FFC000', # Orange - Assessment columns
        'row_alt': 'F2F2F2',           # Light gray - alternating rows
    }

    # Thematic columns definition
    THEMATIC_COLUMNS = {
        'technik': [
            ('AI_Literacies', 'AI Literacies / competences'),
            ('Generative_KI', 'Generative KI (ChatGPT, LLMs, Text-zu-Bild)'),
            ('Prompting', 'Prompt Engineering, Prompt Design'),
            ('KI_Sonstige', 'KI undefiniert / regelbasiert / ML'),
        ],
        'sozial': [
            ('Soziale_Arbeit', 'Social Work, Sozialpädagogik'),
            ('Bias_Ungleichheit', 'Bias / Soziale Ungleichheit / Inequality'),
            ('Gender', 'Geschlecht, Gender Studies'),
            ('Diversitaet', 'Diversity, Inklusion'),
            ('Feministisch', 'Feminismus, feminist theory'),
            ('Fairness', 'Algorithmic fairness, fair AI'),
        ],
        'meta': [
            ('Studientyp', 'Empirisch / Theoretisch / Unclear'),
            ('Decision', 'Include / Exclude / Unclear'),
            ('Exclusion_Reason', 'Grund für Ausschluss'),
            ('Notes', 'Notizen'),
        ]
    }

    def __init__(self, library_id: str, library_type: str = 'group', api_key: str = None):
        """Initialize with Zotero credentials"""
        self.library_id = library_id
        self.library_type = library_type
        self.api_key = api_key

        logger.info(f"Connecting to Zotero {library_type} library: {library_id}")
        self.zot = zotero.Zotero(library_id, library_type, api_key)

        self.items = []
        self.collections = {}

    def fetch_collections(self) -> Dict[str, str]:
        """Fetch all collections and create ID -> Name mapping"""
        logger.info("Fetching collections from Zotero...")
        try:
            all_collections = self.zot.collections()
            for coll in all_collections:
                self.collections[coll['key']] = coll['data']['name']
            logger.info(f"Fetched {len(self.collections)} collections")
            return self.collections
        except Exception as e:
            logger.warning(f"Could not fetch collections: {e}")
            return {}

    def fetch_all_items(self) -> List[Dict]:
        """Fetch all items from Zotero library"""
        logger.info("Fetching items from Zotero API...")

        try:
            all_items = self.zot.everything(self.zot.items())
            logger.info(f"Fetched {len(all_items)} items from Zotero")

            # Filter for publications (not notes, attachments)
            valid_types = ['journalArticle', 'book', 'bookSection', 'conferencePaper',
                          'report', 'thesis', 'webpage', 'document', 'preprint']

            self.items = [item['data'] for item in all_items
                         if item.get('data', {}).get('itemType') in valid_types]

            logger.info(f"Filtered to {len(self.items)} valid publications")
            return self.items

        except Exception as e:
            logger.error(f"Failed to fetch items: {e}")
            raise

    def extract_authors(self, item: Dict) -> str:
        """Extract author names from Zotero item"""
        creators = item.get('creators', [])
        if not creators:
            return ''

        author_names = []
        for creator in creators:
            if creator.get('creatorType') == 'author':
                last_name = creator.get('lastName', '')
                first_name = creator.get('firstName', '')
                if last_name and first_name:
                    author_names.append(f"{last_name}, {first_name[0]}.")
                elif last_name:
                    author_names.append(last_name)
                elif creator.get('name'):
                    author_names.append(creator.get('name'))

        return '; '.join(author_names)

    def normalize_year(self, date_str: str) -> str:
        """Extract year from date string"""
        if not date_str:
            return ''
        year_match = re.search(r'\b(19|20)\d{2}\b', date_str)
        return year_match.group(0) if year_match else ''

    def normalize_text(self, text: str) -> str:
        """Normalize whitespace in text"""
        if not text:
            return ''
        # Replace multiple whitespace with single space
        normalized = re.sub(r'\s+', ' ', text)
        # Remove HTML tags
        normalized = re.sub(r'<[^>]+>', '', normalized)
        return normalized.strip()

    def extract_source_tool(self, collection_names: List[str]) -> str:
        """Extract Deep Research tool from collection names"""
        if not collection_names:
            return 'Manual'

        for coll in collection_names:
            coll_lower = coll.lower()
            if 'claude' in coll_lower:
                return 'Claude'
            elif 'gemini' in coll_lower:
                return 'Gemini'
            elif 'openai' in coll_lower or 'gpt' in coll_lower:
                return 'ChatGPT'
            elif 'perplexity' in coll_lower:
                return 'Perplexity'

        return 'Manual'

    def create_dataframe(self) -> pd.DataFrame:
        """Convert Zotero items to DataFrame"""
        if not self.items:
            logger.warning("No items to convert")
            return None

        logger.info(f"Converting {len(self.items)} items to DataFrame...")

        rows = []
        for idx, item in enumerate(self.items, 1):
            # Extract metadata
            pub_year = self.normalize_year(item.get('date', ''))
            authors = self.extract_authors(item)
            first_author = authors.split(';')[0].split(',')[0] if authors else 'Unknown'

            # Get collection names
            collection_ids = item.get('collections', [])
            collection_names = [self.collections.get(cid, cid) for cid in collection_ids]
            source_tool = self.extract_source_tool(collection_names)

            row = {
                'ID': idx,
                'Zotero_Key': item.get('key', ''),
                'Author_Year': f"{first_author} ({pub_year})",
                'Title': item.get('title', ''),
                'DOI': item.get('DOI', ''),
                'Item_Type': item.get('itemType', ''),
                'Publication_Year': pub_year,
                'Language': item.get('language', ''),
                'Source_Tool': source_tool,
                'Abstract': self.normalize_text(item.get('abstractNote', '')),
                'URL': item.get('url', ''),
            }

            # Add thematic columns (empty)
            for category in ['technik', 'sozial']:
                for col_name, _ in self.THEMATIC_COLUMNS[category]:
                    row[col_name] = ''

            # Add meta columns (empty)
            for col_name, _ in self.THEMATIC_COLUMNS['meta']:
                row[col_name] = ''

            rows.append(row)

        df = pd.DataFrame(rows)

        # Sort alphabetically by Author_Year
        df = df.sort_values('Author_Year', key=lambda x: x.str.lower())
        df['ID'] = range(1, len(df) + 1)  # Renumber after sorting

        logger.info(f"Created DataFrame with {len(df)} rows, {len(df.columns)} columns")
        return df

    def create_excel(self, df: pd.DataFrame, output_file: str):
        """Create formatted Excel file with dropdowns and colors"""
        logger.info(f"Creating Excel file: {output_file}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Assessment"

        # Define styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Color fills
        fill_meta = PatternFill(start_color=self.COLORS['header_meta'],
                                end_color=self.COLORS['header_meta'], fill_type='solid')
        fill_technik = PatternFill(start_color=self.COLORS['header_technik'],
                                   end_color=self.COLORS['header_technik'], fill_type='solid')
        fill_sozial = PatternFill(start_color=self.COLORS['header_sozial'],
                                  end_color=self.COLORS['header_sozial'], fill_type='solid')
        fill_assessment = PatternFill(start_color=self.COLORS['header_assessment'],
                                      end_color=self.COLORS['header_assessment'], fill_type='solid')

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border

                if r_idx == 1:  # Header row
                    cell.font = header_font
                    cell.alignment = header_alignment

                    # Color based on column type
                    col_name = df.columns[c_idx - 1]
                    if col_name in [c[0] for c in self.THEMATIC_COLUMNS['technik']]:
                        cell.fill = fill_technik
                    elif col_name in [c[0] for c in self.THEMATIC_COLUMNS['sozial']]:
                        cell.fill = fill_sozial
                    elif col_name in [c[0] for c in self.THEMATIC_COLUMNS['meta']]:
                        cell.fill = fill_assessment
                    else:
                        cell.fill = fill_meta

        # Set column widths
        column_widths = {
            'A': 5,    # ID
            'B': 12,   # Zotero_Key
            'C': 25,   # Author_Year
            'D': 60,   # Title
            'E': 20,   # DOI
            'F': 15,   # Item_Type
            'G': 8,    # Publication_Year
            'H': 8,    # Language
            'I': 12,   # Source_Tool
            'J': 80,   # Abstract
            'K': 30,   # URL
        }

        # Thematic columns
        col_letter = ord('L')
        for category in ['technik', 'sozial', 'meta']:
            for col_name, _ in self.THEMATIC_COLUMNS[category]:
                column_widths[chr(col_letter)] = 12 if col_name != 'Notes' else 30
                col_letter += 1

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Add data validations (dropdowns)
        num_rows = len(df) + 1

        # Ja/Nein validation for thematic columns
        ja_nein_validation = DataValidation(
            type="list",
            formula1='"Ja,Nein"',
            allow_blank=True
        )
        ja_nein_validation.error = 'Bitte Ja oder Nein auswählen'
        ja_nein_validation.errorTitle = 'Ungültige Eingabe'
        ws.add_data_validation(ja_nein_validation)

        # Find column indices for thematic columns
        col_idx = 12  # Starting after URL (column K = 11)
        for category in ['technik', 'sozial']:
            for col_name, _ in self.THEMATIC_COLUMNS[category]:
                col_letter = chr(ord('A') + col_idx - 1)
                ja_nein_validation.add(f'{col_letter}2:{col_letter}{num_rows}')
                col_idx += 1

        # Studientyp validation
        studientyp_col = chr(ord('A') + col_idx - 1)
        studientyp_validation = DataValidation(
            type="list",
            formula1='"Empirisch,Theoretisch,Unclear"',
            allow_blank=True
        )
        ws.add_data_validation(studientyp_validation)
        studientyp_validation.add(f'{studientyp_col}2:{studientyp_col}{num_rows}')
        col_idx += 1

        # Decision validation
        decision_col = chr(ord('A') + col_idx - 1)
        decision_validation = DataValidation(
            type="list",
            formula1='"Include,Exclude,Unclear"',
            allow_blank=True
        )
        ws.add_data_validation(decision_validation)
        decision_validation.add(f'{decision_col}2:{decision_col}{num_rows}')
        col_idx += 1

        # Exclusion_Reason validation
        exclusion_col = chr(ord('A') + col_idx - 1)
        exclusion_validation = DataValidation(
            type="list",
            formula1='"Not relevant topic,Wrong publication type,Wrong language,Duplicate,No full text,Other"',
            allow_blank=True
        )
        ws.add_data_validation(exclusion_validation)
        exclusion_validation.add(f'{exclusion_col}2:{exclusion_col}{num_rows}')

        # Freeze header row and first columns
        ws.freeze_panes = 'D2'

        # Save
        wb.save(output_file)
        logger.info(f"Excel file saved: {output_file}")
        logger.info(f"  - {len(df)} papers")
        logger.info(f"  - {len(df.columns)} columns")

        return output_file


def main():
    parser = argparse.ArgumentParser(
        description='Create FemPrompt thematic assessment Excel from Zotero'
    )
    parser.add_argument(
        '--library-id',
        default='6080294',
        help='Zotero library ID (default: 6080294 - FemPrompt)'
    )
    parser.add_argument(
        '--library-type',
        default='group',
        choices=['user', 'group'],
        help='Library type (default: group)'
    )
    parser.add_argument(
        '--api-key',
        default=None,
        help='Zotero API key (optional for public groups)'
    )
    parser.add_argument(
        '-o', '--output',
        default='assessment/assessment_full.xlsx',
        help='Output Excel file (default: assessment/assessment_full.xlsx)'
    )

    args = parser.parse_args()

    # Create output directory
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        # Initialize generator
        generator = ThematicAssessmentGenerator(
            library_id=args.library_id,
            library_type=args.library_type,
            api_key=args.api_key
        )

        # Fetch data
        generator.fetch_collections()
        generator.fetch_all_items()

        if not generator.items:
            logger.warning("No items found in Zotero library")
            return

        # Create DataFrame
        df = generator.create_dataframe()

        # Create Excel
        generator.create_excel(df, args.output)

        print("\n" + "="*70)
        print("SUCCESS: Thematic assessment Excel created")
        print("="*70)
        print(f"\nFile: {args.output}")
        print(f"Papers: {len(df)}")
        print(f"\nColumns:")
        print(f"  - Metadata: 11 (ID, Author_Year, Title, etc.)")
        print(f"  - Technik (Ja/Nein): 4 (AI_Literacies, Generative_KI, Prompting, KI_Sonstige)")
        print(f"  - Sozial (Ja/Nein): 6 (Soziale_Arbeit, Bias, Gender, Diversitaet, Feministisch, Fairness)")
        print(f"  - Meta: 4 (Studientyp, Decision, Exclusion_Reason, Notes)")
        print(f"\nNext steps:")
        print(f"  1. Upload to Google Sheets")
        print(f"  2. Share with team for assessment")
        print(f"  3. After assessment: run pipeline for included papers")

    except Exception as e:
        logger.error(f"Error: {e}")
        raise


if __name__ == '__main__':
    main()
